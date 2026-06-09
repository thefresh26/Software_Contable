import io
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser


def _abrir_wb(request):
    import openpyxl
    archivo = request.FILES.get('archivo')
    if not archivo:
        raise ValueError('No se proporcionó archivo.')
    return openpyxl.load_workbook(archivo, read_only=True, data_only=True)


class ImportarTercerosView(APIView):
    parser_classes = [MultiPartParser]

    def post(self, request):
        preview = request.query_params.get('preview') == 'true'
        try:
            wb = _abrir_wb(request)
        except ValueError as e:
            return Response({'error': str(e)}, status=400)

        ws = wb.active
        creados, errores, filas_preview = 0, [], []
        from apps.terceros.models import Tercero

        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue
            try:
                tipo, tipo_persona, nombre, nit, email, telefono, direccion, ciudad = (
                    (row[j] if j < len(row) else None) for j in range(8)
                )
                if not nombre or not nit:
                    raise ValueError('nombre y nit son obligatorios')
                if str(tipo).lower() not in ('cliente', 'proveedor', 'ambos'):
                    tipo = 'cliente'
                if str(tipo_persona).lower() not in ('natural', 'juridica'):
                    tipo_persona = 'natural'
                if preview:
                    filas_preview.append({'fila': i, 'nombre': nombre, 'nit': str(nit), 'tipo': tipo})
                    if len(filas_preview) >= 5:
                        break
                    continue
                Tercero.objects.update_or_create(
                    nit=str(nit).strip(),
                    defaults=dict(
                        tipo=str(tipo).lower().strip(),
                        tipo_persona=str(tipo_persona).lower().strip(),
                        nombre=str(nombre).strip(),
                        email=str(email or '').strip(),
                        telefono=str(telefono or '').strip(),
                        direccion=str(direccion or '').strip(),
                        ciudad=str(ciudad or '').strip(),
                    ),
                )
                creados += 1
            except Exception as e:
                errores.append({'fila': i, 'error': str(e)})

        if preview:
            return Response({'preview': filas_preview})
        return Response({'creados': creados, 'errores': errores})


class ImportarProductosView(APIView):
    parser_classes = [MultiPartParser]

    def post(self, request):
        preview = request.query_params.get('preview') == 'true'
        try:
            wb = _abrir_wb(request)
        except ValueError as e:
            return Response({'error': str(e)}, status=400)

        ws = wb.active
        creados, errores, filas_preview = 0, [], []
        from apps.inventario.models import Producto, Categoria

        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue
            try:
                (codigo, nombre, categoria_nombre, precio_compra, precio_venta,
                 iva_porcentaje, stock_actual, stock_minimo, unidad_medida) = (
                    (row[j] if j < len(row) else None) for j in range(9)
                )
                if not codigo or not nombre:
                    raise ValueError('codigo y nombre son obligatorios')
                if preview:
                    filas_preview.append({'fila': i, 'codigo': codigo, 'nombre': nombre})
                    if len(filas_preview) >= 5:
                        break
                    continue
                categoria = None
                if categoria_nombre:
                    categoria, _ = Categoria.objects.get_or_create(nombre=str(categoria_nombre).strip())
                Producto.objects.update_or_create(
                    codigo=str(codigo).strip(),
                    defaults=dict(
                        nombre=str(nombre).strip(),
                        categoria=categoria,
                        precio_compra=float(precio_compra or 0),
                        precio_venta=float(precio_venta or 0),
                        iva_porcentaje=float(iva_porcentaje or 19),
                        stock_actual=int(stock_actual or 0),
                        stock_minimo=int(stock_minimo or 5),
                        unidad_medida=str(unidad_medida or 'UND').strip(),
                    ),
                )
                creados += 1
            except Exception as e:
                errores.append({'fila': i, 'error': str(e)})

        if preview:
            return Response({'preview': filas_preview})
        return Response({'creados': creados, 'errores': errores})


class ImportarPlanCuentasView(APIView):
    parser_classes = [MultiPartParser]

    def post(self, request):
        preview = request.query_params.get('preview') == 'true'
        try:
            wb = _abrir_wb(request)
        except ValueError as e:
            return Response({'error': str(e)}, status=400)

        ws = wb.active
        creados, errores, filas_preview = 0, [], []
        from apps.contabilidad.models import CuentaPUC

        TIPOS_VALIDOS = ['activo', 'pasivo', 'patrimonio', 'ingreso', 'gasto', 'costo']

        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue
            try:
                codigo, nombre, tipo, nivel = (
                    (row[j] if j < len(row) else None) for j in range(4)
                )
                if not codigo or not nombre:
                    raise ValueError('codigo y nombre son obligatorios')
                tipo = str(tipo or 'activo').lower().strip()
                if tipo not in TIPOS_VALIDOS:
                    raise ValueError(f'tipo inválido: {tipo}')
                if preview:
                    filas_preview.append({'fila': i, 'codigo': codigo, 'nombre': nombre, 'tipo': tipo})
                    if len(filas_preview) >= 5:
                        break
                    continue
                CuentaPUC.objects.update_or_create(
                    codigo=str(codigo).strip(),
                    defaults=dict(
                        nombre=str(nombre).strip(),
                        tipo=tipo,
                        nivel=int(nivel or 4),
                    ),
                )
                creados += 1
            except Exception as e:
                errores.append({'fila': i, 'error': str(e)})

        if preview:
            return Response({'preview': filas_preview})
        return Response({'creados': creados, 'errores': errores})


class PlantillaExcelView(APIView):
    """Genera y descarga una plantilla Excel vacía para cada tipo de importación."""

    PLANTILLAS = {
        'terceros': ['tipo', 'tipo_persona', 'nombre', 'nit', 'email', 'telefono', 'direccion', 'ciudad'],
        'productos': ['codigo', 'nombre', 'categoria', 'precio_compra', 'precio_venta', 'iva_porcentaje', 'stock_actual', 'stock_minimo', 'unidad_medida'],
        'plan-cuentas': ['codigo', 'nombre', 'tipo', 'nivel'],
    }

    def get(self, request, tipo):
        import openpyxl
        columnas = self.PLANTILLAS.get(tipo)
        if not columnas:
            return Response({'error': 'Tipo inválido'}, status=400)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = tipo
        ws.append(columnas)
        # Fila de ejemplo
        ejemplos = {
            'terceros': ['cliente', 'natural', 'Empresa Ejemplo S.A.S', '900123456-7', 'contacto@empresa.com', '3001234567', 'Cra 10 # 20-30', 'Bogotá'],
            'productos': ['PROD001', 'Producto de Ejemplo', 'General', 50000, 75000, 19, 100, 10, 'UND'],
            'plan-cuentas': ['1105', 'Caja', 'activo', 4],
        }
        ws.append(ejemplos.get(tipo, []))
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        response = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="plantilla_{tipo}.xlsx"'
        return response
