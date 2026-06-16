import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse

_HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
_HEADER_FILL = PatternFill(fill_type='solid', fgColor='1E3A5F')
_TITLE_FONT = Font(bold=True, size=13)
_CENTER = Alignment(horizontal='center')
_RIGHT = Alignment(horizontal='right')
CONTENT_TYPE = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'


def excel_response(wb, filename):
    # HttpResponse doesn't support tell()/seek() which zipfile requires;
    # write to BytesIO first then pass raw bytes to HttpResponse.
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type=CONTENT_TYPE)
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def nuevo_libro(titulo_hoja, empresa_nombre, subtitulo, cabeceras):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = titulo_hoja
    n = len(cabeceras)
    col_fin = chr(64 + n) if n <= 26 else 'Z'

    ws.merge_cells(f'A1:{col_fin}1')
    ws['A1'] = empresa_nombre
    ws['A1'].font = _TITLE_FONT
    ws['A1'].alignment = _CENTER

    ws.merge_cells(f'A2:{col_fin}2')
    ws['A2'] = subtitulo
    ws['A2'].alignment = _CENTER

    fila_cab = 4
    for i, col in enumerate(cabeceras, 1):
        cell = ws.cell(row=fila_cab, column=i, value=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER

    return wb, ws, fila_cab + 1


def ajustar_columnas(ws):
    for col in ws.columns:
        try:
            widths = []
            for c in col:
                try:
                    widths.append(len(str(c.value)) if c.value is not None else 0)
                except AttributeError:
                    pass
            ancho = max(widths, default=10)
            first = col[0]
            ws.column_dimensions[first.column_letter].width = min(ancho + 4, 55)
        except (IndexError, AttributeError, TypeError):
            pass
